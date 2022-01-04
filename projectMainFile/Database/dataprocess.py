import os
import sqlite3
import openpyxl
import subprocess

def sqlprompt(row, type):
    result = '''
    insert into course values (DEFAULT,'''
    for text in row:
        try:
            text = text.replace(" ", "")
        except:
            continue;
    clsromlst = [12,13,18,19,20,21]
    clasromstr = [""]*6
    for index in clsromlst:
        if (len(row[index]) > 3):
            clasromstr.append(row[index])
    for i in range(0,12):
        try:
            temp = int(row[i])
            result += str(temp) + ","
        except:
            result += "\"" + row[i] + "\"" + ","
    for i in range(0, 2):
        result +=  "\"" + clasromstr[i] + "\"" + ","
    for i in range(14, 18):
        try:
            temp = int(row[i])
            result += str(temp) + ","
        except:
            result += "\"" + row[i] + "\"" + ","
    for i in range(2, 6):
        result +=  "\"" + clasromstr[i] + "\"" + ","
    for i in range(21, 26):
        try:
            temp = int(row[i])
            result += str(temp) + ","
        except:
            result += "\"" + row[i] + "\"" + ","
    timeList = getTimeList(row[15])
    for i in range(0,20):
        result += str(timeList[i]) + ","
    if "限大一" in row[16]:
        result += "1,"
        if "限大二" in row[16]:
            result += "2,"
            if "限大三" in row[16]:
                    result += "3,"
                    if "限大四" in row[16]:
                        result += "4,"
                    else:
                        result+= "0,"
    if "限本系" in row[16]:
        result += "true,"
    else:
        result += "false,"
    result += type[0] + ","
    if type[0] == '2' and "通識" in row[16]:
        result += '1' + ","
    else:
        result += '-1' + ","
    if "初選不開放" in row[16]:
        result += "true,"
    else:
        result += "false,"
    libcat = []
    if "文學與藝術" in row[16]:
        libcat.append(1)
    if "歷史思維" in row[16]:
        libcat.append(2)
    if "世界文明" in row[16]:
        libcat.append(3)
    if "哲學與道德" in row[16]:
        libcat.append(4)
    if "公民意識" in row[16]:
        libcat.append(5)
    if "量化分析" in row[16]:
        libcat.append(6)
    if "物質科學" in row[16]:
        libcat.append(7)
    if "生命科學" in row[16]:
        libcat.append(8)
    libcat.append(0)
    libcat.append(0)
    result += f"{libcat[0]}, {libcat[1]}"
    result += ");"
    return result

def getTimeList(str):
    dir = os.getcwd()+"\\a"
    return subprocess.Popen(dir + " " + str, stdout=subprocess.PIPE).communicate()[0].split()

dir = os.getcwd()
db = sqlite3.connect(dir + 'course.db');
cursor = db.cursor()
workbook = openpyxl.load_workbook(dir + "\\CourseExcel\\" + input("input the name of the excel file"))
sheet = workbook.worksheets[0]
type =  [input("Category: 0 共同 1 通識 2 系上課"), input("Category2: 0 共同 1 通識 2 系上課")]

#https://ithelp.ithome.com.tw/articles/10246377
for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row, values_only=True):
    print(sqlprompt(row, type))
    cursor.execute(sqlprompt(row, type))
db.commit()
